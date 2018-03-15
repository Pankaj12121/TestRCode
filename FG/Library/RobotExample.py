from openpyxl import load_workbook
from openpyxl import Workbook
import xlrd
import os
from sys import getsizeof
import os, time,datetime
import shutil
from datetime import datetime,timedelta
import copy
import logging

class RobotExample:
    #config_file_path = 'C:\Projects\FG\ConfigurationSheet.xlsx'
    #output_folder_to_save_macro_file = 'C:\\New folder\\MacroFiles'
    #macro_file_path = 'C:\\Projects\\Copy of MemberUploadFile12354.xlsm'
    #time_stamp_folder_path = 'C:\\New folder\\tempfolder'
    #default_directory = 'C:\\Projects\\New folder'
    def __init__(self):
        #logger = logging.getLogger('myapp')
        print '\n'
    def xlsmFileWriting(self,masterFilePath):        
        #wb = load_workbook('C:\Projects\Copy of MemberUploadFile12354.xlsm', keep_vba=True)
        wb = load_workbook('masterFilePath', keep_vba=True)
        print(wb.sheetnames)
        ws = wb["Sheet1"]
        ws.cell(row=10, column=2).value = "Updated"
        wb.save('masterFilePath')
    def get_xl_values(self,filepath,sheetName=None):
        """Returns the dictionary of values given row in the MS Excel file """
        workbook = xlrd.open_workbook(filepath)
        snames=workbook.sheet_names()
        if sheetName==None:
            sheetName=snames[0]
        worksheet=workbook.sheet_by_name(sheetName)
        noofrows=worksheet.nrows
        noOfColmns = worksheet.ncols
        print 'no of rows:',noofrows
        print 'no of colmns: ',noOfColmns
        return 'no of rows:',noofrows
        return 'no of colmns: ',noOfColmns
    def sample(self):
        src_file = 'C:\\Projects\\FG\\Project Code\\FGE\\sample.xls'
        file_names = src_file.split('\\')
        len_file = len(file_names)
        print 'len_file',len_file
        file_name = file_names[len_file-1]
        print 'file_name',file_name
##        for x in range (1,4):            
##            LOG_FILENAME = 'C:\\Projects\\FG\\Project Code\\FGE\\sample'+str(x)+'.out'
##            logging.basicConfig(filename=LOG_FILENAME,
##                        level=logging.DEBUG,
##                        )
##            logger = logging.getLogger('Logging Data')
##            logger.error('Number')

    def get_list_of_excel_files_from_directory(self,default_directory):
        try:
            xl_file_list = []
            extension = '.xls'
            xlsxext = '.xlsx'
            list_dir = os.listdir(default_directory)
            print 'default folder path:',default_directory
            print 'files in default folder',list_dir
            listlen = len(list_dir)
            if listlen == 0:zczsfssaf
                return False,'No files available in Directiory: '+default_directory+'. Please check'
                #raise ValueError('No files available with in Directiory:',default_directory)
            for listf in list_dir:
                if listf.endswith(extension) or listf.endswith(xlsxext):
                    xl_file_list.append(default_directory+'\\'+listf)
            print 'file list with defined extensions',xl_file_list
            print len(xl_file_list)
            if len(xl_file_list) == 0:
                return False, "No files available in default directory "+default_directory+" with extensions "+extension+" and "+xlsxext
            return True,xl_file_list
##        
##            for efile in xl_file_list:
##                log_folder_with_time_stamp = ''
##                fileName = efile.replace(default_directory+'\\','')
##                policyNumber = fileName.split('.')[0]
##                print 'Policy Number : ',policyNumber
##
##                error_log_file_path =self.get_error_log_file_path_for_policy_number(outpu_folder_for_log_errors,policyNumber)
##                            
##                status,sheet_name,err_msg = self.consolidate_excel_file(efile,config_file_path,output_folder_to_save_macro_file,macro_file_path,policyNumber,lookupdata_file_path)
##                if status == False:
##                    print 'self.error_log_file_path',self.error_log_file_path
##                    print 'output_folder_to_move_error_files',output_folder_to_move_error_files
##                    print 'efile',efile
##                    self.move_selected_file_to_required_folder(efile,output_folder_to_move_error_files)
##                    self.write_error_msg_to_excel_sheet(error_log_file_path, policyNumber,sheet_name,status,err_msg)
##                else:
##                    self.move_selected_file_to_required_folder(efile,output_folder_for_successfully_consolidated_fies)
##            return True,'Pass'
        except Exception, e:
            print 'Exception in getting the file list from client data folder. ERROR: '+str(e)
            #self.write_error_msg_to_excel_sheet(self.error_log_file_path, policyNumber,'','False','Exception in reading the client file. ERROR: '+str(e))
            return False, 'File exception in read client data files and writing to macro files from default folder. ERROR: '+str(e)
            #raise ValueError ('File exception in read client data files and write to macro files from default folder')
            
    
    def get_error_log_file_path_for_policy_number(self,outpu_folder_for_log_errors,policyNumber,time_stamp):
        #time_stamp = time.strftime("%d-%m-%y-%H-%M-%S")
        log_folder_with_time_stamp = outpu_folder_for_log_errors+'\\'+str(time_stamp)
##        if not os.path.exists(log_folder_with_time_stamp):
##            os.makedirs(log_folder_with_time_stamp)                
        error_log_file_path = log_folder_with_time_stamp+'\\Error_Log_For_'+policyNumber+'.xlsx'
        return error_log_file_path,log_folder_with_time_stamp

    
    def consolidate_client_excel_file(self,filePath,config_file_path,output_folder_to_save_macro_file,macro_file_path,policyNumber,lookupdata_file_path, error_log_file_path,folder_path):

        try:
            
            workbook = xlrd.open_workbook(filePath)
            sheet_names=workbook.sheet_names()
            print sheet_names,'\n'
            master_file_dict, status, err_msg = self.get_ms_excel_multiple_row_values_into_dictionary_list_based_on_key(lookupdata_file_path,policyNumber)
            if status == 'False':
                return False,'',err_msg
            print 'master_file_dict\n',master_file_dict
            err_dict = {}
            len_status = True
            s_n = ''
            len_err_msg = ''
            sheet_data_dict = {}
            for sheet_name in sheet_names:            
                ''' Get default sheet name for the different sheet names expectd from theconfigurations sheet'''
                defaultSheetName,status,err_msg = self.get_default_sheet_name_from_config_sheet(sheet_name,config_file_path)
                if status == 'False':
                    status ,msg = self.write_error_msg_to_excel_sheet(folder_path,error_log_file_path, policyNumber,sheet_name,'Get Default Sheet Name From Config File',status,err_msg)
                    continue
                    #return False,sheet_name,err_msg

                '''get the values in the sheet in dictionary format with key as header name and value as the column value in multiple dictionary'''
                ''' sheetDict format = {1={col1=val1,col2=val2},2={col1=val3,col2=val4}}'''
                sheet_data, status, err_msg = self.get_sheet_data_from_client_file(workbook,sheet_name,defaultSheetName,config_file_path,lookupdata_file_path,policyNumber,master_file_dict,folder_path, error_log_file_path)
                print 'sheet_data, status, err_msg\n',sheet_data, status, err_msg                
                if status == 'False':
                    status ,msg = self.write_error_msg_to_excel_sheet(folder_path,error_log_file_path, policyNumber,sheet_name,'Get Sheet Data',status,err_msg)
                    continue
                    #return False, sheet_name, err_msg
                print 'status, err_msg',status, err_msg
                print 'complete sheet dictionary:\n',sheet_data,'\n\n'
                dictlen = len(sheet_data)
                if dictlen == 0:
                    len_err_msg = 'No values available in the sheet/No config sheet value provided to map the headers for the sheet: '+sheet_name
                    status ,msg = self.write_error_msg_to_excel_sheet(folder_path,error_log_file_path, policyNumber,sheet_name,'Get Sheet Data','False',len_err_msg)
                    continue
                sheet_data_dict[str(defaultSheetName)] = sheet_data
                 
            
            for sheet_name in sheet_data_dict:
                print '====================================================================\n'
                print 'sheet_data_dict\n',sheet_data_dict,'\n'
                print '======================================================================'
                date = time.strftime("%d-%m-%Y")
                policyFolderPath = output_folder_to_save_macro_file+'\\'+policyNumber
                file_path_to_save = policyFolderPath+'\\'+policyNumber+'_Upload('+sheet_name+')'+date+'.xlsm'
                print 'Writing data to to: ',file_path_to_save
                ind_sheet_data = sheet_data_dict[str(sheet_name)]
                status, err_msg = self.write_sheet_data_to_macro_file(macro_file_path,ind_sheet_data,file_path_to_save,policyFolderPath)
                if status == 'False':
                    status ,msg = self.write_error_msg_to_excel_sheet(folder_path,error_log_file_path, policyNumber,sheet_name,'Write Sheet Data','False',err_msg)
                    continue
                    #return False, sheet_name, err_msg
                print 'sheet_name\n',sheet_name
                    
            return True,'','Pass'
        except Exception, e:
            print 'Exception in reading file'+filePath+' ',str(e)
            return False,'','Exception in reading file'+filePath+'. ERROR:'+str(e)
            #raise ValueError ('File exception in fg script failed to read data from given files',filePath)
            
            
            


    def get_sheet_data_from_client_file(self,workbook,sname,defaultSheetName,config_file_path,lookupdata_file_path,policyNumber,master_file_dict,folder_path, error_log_file_path):

        try:
            no_of_days_to_add, status, err_msg = self.get_number_of_days_to_add(config_file_path)
            if status == 'False':
                return 'None',status, err_msg

            worksheet = workbook.sheet_by_name(sname)

            ''' Get the Header Row number by using the column names provided in config sheet'''
            rowNumber, status, err_msg = self.get_row_number_for_column_name(worksheet,defaultSheetName,config_file_path)
            print 'rowNumber',rowNumber
            if status == 'False':
                return 'None', status, err_msg
            print 'sheet name :' ,sname,'row number: ', rowNumber,'\n'
            noOfRows = worksheet.nrows
            noOfCols = worksheet.ncols

            if rowNumber+1 == noOfRows:
                print 'rowNumber+1 == noOfRows',rowNumber+1, noOfRows
                return 'None','False','No data available in sheet: '+str(defaultSheetName)+'. Header row number is :'+str(rowNumber+1)+' and number of rows in the sheet is: '+str(noOfRows)

            ''' To identify the Type and cover basis based on the policy data in the master data file'''
            policy_master_row_dict,status, err_msg = self.get_ms_excel_row_values_into_dictionary_based_on_key(lookupdata_file_path,policyNumber)
            if status == 'False':
                return 'None',status, err_msg
            cover_base = int(3)
            print 'policy_master_row_dict',policy_master_row_dict
            grade_status = 'False'
            if 'Cover Basis' in policy_master_row_dict.keys() and defaultSheetName != 'Deletion':
                cover_base = policy_master_row_dict['Cover Basis']
                if int(cover_base) == 3:
                    if 'Type' in policy_master_row_dict: 
                        type_of_structure = policy_master_row_dict['Type']
                        if type_of_structure == 'Grade':
                            grade_status = 'True'
                    else:
                       return  'None','False','Type is not provided for the policy in master data file. Please add the type in Mater data'
            
            if defaultSheetName == 'Additions':
                if 'Cover Basis' in policy_master_row_dict.keys():
                    cover_base = policy_master_row_dict['Cover Basis']
                    cover_base = int(cover_base)
                else:
                    return 'None','False','Cover basis is not available in the Master data file for the Poilicy Number: '+str(policyNumber)
             
            ''' get the client file sheet header columns in dictionary with header as key and column index as value using the row number'''
            headerDict ,status, err_msg= self.get_client_file_header_values_into_dict(worksheet,rowNumber,defaultSheetName)
            if status == 'False':
                return 'None',status, err_msg
            print 'Default Colmns before mapping Dict in client file, sheet name: ',sname,'\n',headerDict,'\n'
            noOfCols = len(headerDict)

            '''get the mapped headers in a dictionary for the headerdict from above and set the dictionary key as column index and value as default column name'''
            mapped_col_dict ,status, err_msg = self.get_header_mapping_names_from_config_sheet(config_file_path,defaultSheetName,headerDict)
            if status == 'False':
                return 'None',status, err_msg
            print 'mapped Dict: \n',mapped_col_dict,'\n'

            ''' If there are no column mapping values provided then no macro file will created for that'''
            if len(mapped_col_dict) == 0:
                print 'No Column mapping values available for sheet:',defaultSheetName
                return mapped_col_dict,'False','No Mapped columns available for the sheet: '+defaultSheetName
            count = 1
            colValueDict = {}
            for key in mapped_col_dict:
                value = mapped_col_dict[key]
                colValueDict[str(value)] = str(count)
                count = count+1
            print 'colValueDict:\n',colValueDict
            if grade_status == 'True':
                if not 'Band' in colValueDict.keys():
                    err_msg = 'For the cover base 3 with Grade type policy, no expected columns are provided for the Band in config file for sheet:'+defaultSheetName+' or no column available for Band/Grade'
                    return 'None','False',err_msg
                    

            status,err_msg = self.verify_mandatory_fields_availability(colValueDict,defaultSheetName)
            if status == 'False':
                err_msg = err_msg+'. Expected column mapping is not provided in config file for the mandatory column'
                return 'None','False',err_msg
            
            ''' verify the Salry related column availability in Additions sheet if cover basis is 2 '''
            if cover_base == 2 and defaultSheetName == 'Additions':
                sal_col_status = 'false'
                for noc in mapped_col_dict:
                    col_name = mapped_col_dict[str(noc)]
                    if col_name == 'Salary /Earning':
                        sal_col_status = 'true'
                        break
                if sal_col_status == 'false':
                    return 'None','False','Salary data is not available for the Addition sheet or expected column for Salary/Earning is not found in config file Additions sheet'
            sheet_dict = {}
            size = int(0);

            male_list = ['m','male','mr','mr.']
            fe_male_list = ['f','fm','female','fe-male','ms','ms.']
            subsidiary_status = 'False'
            
            for nor in range(rowNumber+1, noOfRows):
                try:
                        
                    dictVar = {}
                    size = size+1
                    subsidary_name = ''
                    band_name = ''
                    row_err_msg = ''
                    row_status = 'true'
                    for noc in mapped_col_dict:
                        try:  
                            '''Get the column name from the mapped_col_dict using the column index and value from the sheet'''
                            macro_file_column_name = mapped_col_dict[str(noc)]                        
                            
                            column_value_in_client = worksheet.cell_value(nor,int(noc))
                            if column_value_in_client == '':
                                continue
                            print 'column_value_in_client before other actions,',column_value_in_client
                            if (macro_file_column_name == 'Subsidiary'):
                                subsidiary_status = 'True'
                                subsidary_name = column_value_in_client
                                print 'subsidary_name',subsidary_name
                                #column_value_in_client = self.get_subsidiary_data_into_dictonary(macro_file_column_name,column_value_in_client,master_file_dict)
                                continue

                            ''' get the gender, salutation and surname based on gender value in client''' 
                            if macro_file_column_name=='Gender':
                                print '\nEntered in Gender:  Gender Value:\n',column_value_in_client
                                
                                if type(column_value_in_client) is unicode:
                                    print '\n============================================\nEntered in unicode\n=============================================\n'
                                    column_value_in_client = column_value_in_client.encode('utf-8')
                                print '\ntype(column_value_in_client)\n',type(column_value_in_client)
                                print '\ntype(column_value_in_client) is str\n',(type(column_value_in_client) is str)
                                if not type(column_value_in_client) is str:
                                    row_status = 'false'
                                    row_err_msg = row_err_msg+'Gender value:'+str(column_value_in_client)+' is not string type. Please check the data in row: '+str(int(nor)+1)+' and col: '+str(int(noc)+1)+'. \n'
                                    #status ,msg = self.write_error_msg_to_excel_sheet(folder_path, error_log_file_path, policyNumber,defaultSheetName,'Reading Sheet Row Data','False',err_msg)
                                    #continue
                                else:                             
                                    column_value_in_client = column_value_in_client.lower()
                                    if column_value_in_client in male_list:
                                        column_value_in_client = 'M'
                                    if column_value_in_client in fe_male_list:
                                        column_value_in_client = 'F'
                                    if column_value_in_client == 'M':
                                        dictVar['Salutation'] = 'MR'
                                        dictVar['SurName'] = 'MR'
                                    if column_value_in_client == 'F':
                                        dictVar['Salutation'] = 'MS'
                                        dictVar['SurName'] = 'MS'
                                
                            '''Convert the dates into the required date format into macro file'''
                            if macro_file_column_name == 'Effective\n Date' or macro_file_column_name == 'Date of\nJoining' or macro_file_column_name == 'DOB':
                                for col_name in headerDict:
                                    if str(headerDict[col_name]) == str(noc):
                                        print 'headerDict[col_name] == noc:\n\n',headerDict[col_name], noc,'\n\n'
                                        client_column_name = headerDict[col_name]
                                        print 'client_column_name ',col_name
                                        break
                                '''Change the date value from float to date format'''
                                print 'col_name, column_value_in_client, macro_file_column_name\n',col_name, column_value_in_client, macro_file_column_name
                                column_value_in_client ,status, err_msg= self.format_date_to_dd_mm_yyyy(defaultSheetName,col_name, column_value_in_client, macro_file_column_name,no_of_days_to_add,workbook)
                                print 'column_value_in_client for date ,status, err_msg\n',column_value_in_client ,status, err_msg
                                if status == 'False':
                                    row_status = 'false'
                                    row_err_msg = row_err_msg+'. '+err_msg+'. In row: '+str(int(nor)+1)+' and col: '+str(int(noc)+1)+'. \n'
                                    #status ,msg = self.write_error_msg_to_excel_sheet(folder_path, error_log_file_path, policyNumber,defaultSheetName,'Reading Sheet Row Data','False',err_msg)
                                    #continue
                                    #return 'None',status, err_msg
                            print 'column_value_in_client after all operations',column_value_in_client
                            dictVar[str(macro_file_column_name)] = column_value_in_client
                        except Exception,e:
                            row_err_msg = row_err_msg+'. Exception in reading the row: '+str(int(nor)+1)+' and col: '+str(int(noc)+1)+'. ERROR: '+str(e)
                            status ,msg = self.write_error_msg_to_excel_sheet(folder_path, error_log_file_path, policyNumber,defaultSheetName,'Reading Sheet Row Data','False',row_err_msg)

                    print 'dictVar for row', nor,' Dict:\n',dictVar
                    if len(dictVar) == 0:
                        print 'len(dictVar',len(dictVar)
                        err_msg = 'No data available in the sheet: '+defaultSheetName+' and in row: '+str(int(nor)+1)
                        status ,msg = self.write_error_msg_to_excel_sheet(folder_path, error_log_file_path, policyNumber,defaultSheetName,'Reading Sheet Row Data','False',err_msg)
                        continue
                    emp_code = ''
                    if 'Emp\nNo' in dictVar.keys():
                        emp_code = dictVar['Emp\nNo']
                    else:
                        row_err_msg = 'Employee Number is not available in row: '+str(int(nor)+1)
                        status ,msg = self.write_error_msg_to_excel_sheet(folder_path, error_log_file_path, policyNumber,defaultSheetName,'Reading Sheet Row Data','False',row_err_msg)
                        continue
                    status,err_msg = self.verify_mandatory_fields_availability(dictVar,defaultSheetName)
                    if status == 'False':
                        row_status = 'false'
                        row_err_msg = row_err_msg+'. '+err_msg+' and in row: '+str(int(nor)+1)
                    ''' Verify the row_status if any of the col value failed then log the related error message '''
                    if row_status == 'false':
                        status ,msg = self.write_error_msg_to_excel_sheet(folder_path, error_log_file_path, policyNumber,defaultSheetName,'Reading Sheet Row Data','False',row_err_msg,emp_code)
                        continue

                    ''' Get the band name to verify that with master file strcture name and get the master data for that row '''
                    if grade_status == 'True':
                        if not 'Band' in colValueDict.keys():
                            err_msg = 'No data available for Band/Grade column for the Graded policy type(from master data)'
                            status ,msg = self.write_error_msg_to_excel_sheet(folder_path, error_log_file_path, policyNumber,defaultSheetName,'Reading Sheet Row Data','False',err_msg,emp_code)
                            continue
                    if 'Band' in dictVar.keys():
                        band_name = dictVar['Band']
                        print 'band_name&&&&&&&&&&&&&&&&&&&&&&&&',band_name

                    ''' If cover base is 2 then client should provide the salary value otherwise fail'''
                    if defaultSheetName == 'Additions' and cover_base == 2:
                        if not ('Salary /Earning' in dictVar.keys()):
                            print 'Salary /Earning is not in dictionary\n'
                            err_msg = 'Salary data is not available for the policy '+policyNumber+' in '+defaultSheetName+' sheet row no: '+str(int(nor)+1)+' which have cover base as 2.'
                            status ,msg = self.write_error_msg_to_excel_sheet(folder_path, error_log_file_path, policyNumber,defaultSheetName,'Reading Sheet Row Data','False',err_msg,emp_code)
                            continue
                            #return dictVar,'False',err_msg
                    
                    ''' Mapping of the master data and adding it to the row dictionsry and only for Additions and Revision sheets '''
                    if defaultSheetName == 'Additions' or defaultSheetName == 'Revisions':
                        
                        master_file_mapping_dict,status, err_msg = self.mapping_master_file_data_for_plan_and_occp(master_file_dict,policyNumber,subsidary_name,cover_base,grade_status,subsidiary_status,band_name)
                        if status == 'False':
                            err_msg = err_msg+'. In row: '+str(int(nor)+1)
                            status ,msg = self.write_error_msg_to_excel_sheet(folder_path, error_log_file_path, policyNumber,defaultSheetName,'Reading Sheet Row Data','False',err_msg,emp_code)
                            continue
                            #return 'None',status, err_msg
                        print 'master_file_dict',master_file_mapping_dict
    ##                    if master_file_mapping_dict == 'None':
    ##                        return dictVar,'False',msg+' and sheet name: '+defaultSheetName
                        for key in master_file_mapping_dict:                            
                            if key == 'OCCp' or key== 'Z' or key == 'City':
                                if defaultSheetName == 'Revisions':
                                    continue
                            value = master_file_mapping_dict[key]
                            dictVar[str(key)]=value

                    static_Values ,status, err_msg=self.get_static_values_into_dictionary(config_file_path,defaultSheetName)
                    if status == 'False':
                        err_msg = err_msg+'. In row: '+str(int(nor)+1)
                        status ,msg = self.write_error_msg_to_excel_sheet(folder_path, error_log_file_path, policyNumber,defaultSheetName,'Reading Sheet Row Data','False',err_msg,emp_code)
                        continue
                        #return 'None',status, err_msg
                    for key in static_Values:
                        if key == 'Marrried' or key == 'Bus/Res' or key == 'Nationality':
                            if defaultSheetName != 'Additions':
                                continue
                        if key == 'Termination Reason' and defaultSheetName != 'Deletion':
                            continue
                        value= static_Values[key]
                        dictVar[str(key)] = value
                    if defaultSheetName == 'Additions':
                        if not 'Gender' in dictVar.keys():
                            dictVar['Gender'] = 'M'
                            dictVar['Salutation'] = 'MR'
                            dictVar['SurName'] = 'MR'
                        if not 'Given Name' in dictVar.keys():
                            dictVar['Given Name'] = dictVar['Emp\nNo']
                    if not 'Date of\nJoining' in dictVar.keys():
                        dictVar['Date of\nJoining'] = dictVar['Effective\n Date']
                    dictVar['Policy'] = str(policyNumber)
                    sheet_dict[size] = dictVar
                except Exception, e:
                    err_msg = 'Exception in reading the row: '+str(int(nor)+1)+' in sheet: '+defaultSheetName+'. ERROR: '+str(e)
                    status ,msg = self.write_error_msg_to_excel_sheet(folder_path, error_log_file_path, policyNumber,defaultSheetName,'Reading Sheet Row Data','False',err_msg,emp_code)
            return sheet_dict,'True','Pass'
        except Exception, e:
            print 'Not able to read or write the data from client data Error: ',e
            #raise ValueError ('File exception in get sheet data from client file keyword read the data from different files',workbook)
            return 'None','False','Not able to read the data from the sheet: '+str(defaultSheetName)+' ERROR: '+str(e)
    def verify_mandatory_fields_availability(self,dictVar,defaultSheetName):
        err_msg = ''
        status = 'true'
        if not 'Effective\n Date' in dictVar.keys():
            status = 'false'
            err_msg = err_msg+'Effective\n Date is not available in sheet: '+str(defaultSheetName)
        if (not 'DOB' in dictVar.keys()) and defaultSheetName == 'Additions':
            status = 'false'
            err_msg = err_msg+'DOB is not available in sheet'+str(defaultSheetName)
        if not 'Emp\nNo' in dictVar.keys():
            status = 'false'
            err_msg = err_msg+'Emp No is not available in sheet'+str(defaultSheetName)
        if status == 'false':
            return 'False', err_msg
        return 'True','Pass'
        get_ms_excel_multiple_row_values_into_dictionary_list_based_on_key_in_any_col
    def get_number_of_days_to_add(self,config_file_path):
        try:
            
            config_workbook = xlrd.open_workbook(config_file_path)
            config_worksheet = config_workbook.sheet_by_name('Deletion')
            value = config_worksheet.cell_value(4,2)
            value = int(value)
            print 'Value',value
            return value,'True','Pass'
        except Exception, e:
            return 'None','False','Not able to read the data from configuration file ''Deletion'' sheet. ERROR: '+str(e)

    def get_row_number_for_column_name(self, client_worksheet,defaultSheetName,config_file_path):
        try:

            config_workbook = xlrd.open_workbook(config_file_path)
            config_worksheet = config_workbook.sheet_by_name('SheetNames')
            config_noOfRows = config_worksheet.nrows
            print 'config_noOfRows: ',config_noOfRows
            sheetNamedictVar = {}
            for row in range(1, config_noOfRows):
                sheet_name= config_worksheet.cell_value(row,1)
                column_name_to_get_row_number = config_worksheet.cell_value(row,2)
                sheetNamedictVar[str(sheet_name)] = str(column_name_to_get_row_number)

            colName = sheetNamedictVar[defaultSheetName]
            
            noOfRows = client_worksheet.nrows
            noOfCols = client_worksheet.ncols

            expected_col_names = colName.split(",")
            for colName in expected_col_names:            
                for col in range(0, noOfCols):
                    for row in range(0, noOfRows):
                        cellval=client_worksheet.cell_value(row,col)
                        if str(cellval).lower()== str(colName).lower():
                            return row,'True','Pass'
            print 'No row number'
            return 'None','False','Expected_col_names are not available in client file to identify Header Row Number for the sheet: '+defaultSheetName+'. Expected Cols to identify Header row: '+str(expected_col_names)+'. Please add the expected header for sheet in config file SheetNames sheet ''ColumnNameToGetRowNumber'' column'
            #raise ValueError ('Expected_col_names are not available in client file to identify Header Row Number')
        except Exception, e:
            print 'Exception in getting the row number using the column name. Sheet name:',defaultSheetName,' ERROR:',e
            return 'None','False','Exception in identifying the header row number based on the provided expected header names in configuration file sheet: SheetNames. For the sheet:'+defaultSheetName+' ERROR: '+str(e)
            #raise ValueError ('File exception in get row number for column name keyword from different sheets', client_worksheet)

    def get_client_file_header_values_into_dict(self,worksheet,rowNumber,defaultSheetName):
        try:
            """Returns the list of values given row in the MS Excel file """
            noofrows=worksheet.nrows
            tempList=[]
            row=worksheet.row(rowNumber)
            headerDict = {}
            companyorSubsidary = 'Company/ Subsidary Name'
            for colno in range(0,len(row)):
                cellval=worksheet.cell_value(rowNumber,colno)
                if defaultSheetName=='Transfer' and cellval == companyorSubsidary:
                    cellval = self.append_additional_col(cellval,rowNumber,colno,worksheet)
                print 'cell value : ',cellval,'\n'
                if cellval.strip() == '':
                    continue
                headerDict[str(cellval)] = colno
                #tempList.append(cellval)
            return headerDict,'True','Pass'
        except Exception, e:
            print 'Exception in reading the headers from the client file for sheet name:',defaultSheetName,' ERROR: ',e
            return 'None','False','Exception in raeding the client file header value in sheet:'+str(defaultSheetName)+' in row number:'+str(rowNumber)+' ERROR: '+str(e)
            #raise ValueError('get client file header values in to dictionary',worksheet)
        
    def append_additional_col(self,cellval,rowNumber,colno,worksheet):
        try:
            rowNumber = rowNumber-1
            newVal=worksheet.cell_value(rowNumber,colno)
            newstr = " ".join((newVal,cellval))
            return newstr
        except Exception, e:
            print 'Exception in adding the multiple column names in Transfer sheet for Transfer In andTransfer out Company/Subsidary. ERROR: ',e
            return False
            #raise ValueError('File exception in append additional columns',worksheet)
        
            
        
    def get_header_mapping_names_from_config_sheet(self,config_file_path,defaultSheetName,headerDict):
        try:
            workbook = xlrd.open_workbook(config_file_path)
            worksheet = workbook.sheet_by_name(defaultSheetName)
            noOfRows = worksheet.nrows
            columnMappingdictVar = {}
            
            for row in range(1, noOfRows):
                cellval= worksheet.cell_value(row,0)
                
                defaultValue = worksheet.cell_value(row,1)
                columnMappingdictVar[str(defaultValue)] = str(cellval)
        
            print 'columnMappingdictVar\n',columnMappingdictVar,'\n'
            dictLen = len(columnMappingdictVar)
            if dictLen == 0:
                return columnMappingdictVar,'False','No Column mapping values provided in confiuration file for the sheet: '+str(defaultSheetName)
            defaultColDict = {}
            for header in headerDict:
                #status = 'false'
                
                colIndex = headerDict[header]
                #header = header.lower()
                #print 'Header after lower',header
##                if dictLen == 0:
##                    defaultColDict[str(colIndex)]=header
##                    continue
                for defaultColumn in columnMappingdictVar:
                    expectedColList = columnMappingdictVar[defaultColumn]
                    expectedColList = expectedColList.split(",")
                    sal = copy.copy(expectedColList)
                    for name in sal:
                        expectedColList.append(name.lower().replace(" ", ""))
                    
                    ''' Convert header from client file to lower case and replace the spaces with empty finally we will verify the header which is lower and have no space in it'''
                    ''' Expected col list dictionary contains the expected values provided in config sheet and lower case values of that and removed spaces'''
                    if header.lower().replace(" ", "") in expectedColList:
                        defaultColDict[str(colIndex)] = str(defaultColumn)
                        status = 'true'
                        continue
##                if status == 'false':
##                    defaultColDict[str(colIndex)] = str(header)
                    
            return defaultColDict,'True','Pass'
        except Exception, e:
            print 'Exception in Header Mapping for sheet: ',defaultSheetName,' ERROR: ',str(e)
            return 'None','False','Exception in mapping the headers in client file sheet: '+str(defaultSheetName)+' with macro file headers. ERROR: '+str(e)
            #raise ValueError ('File exception in get header mapping names from configaration sheet',config_file_path)
        

            

    def get_default_sheet_name_from_config_sheet(self,sname,config_file_path):
        try:
            workbook = xlrd.open_workbook(config_file_path)
            worksheet = workbook.sheet_by_name('SheetNames')
            noOfRows = worksheet.nrows
            sheetNamedictVar = {}

            '''get the default sheet names mapped with the expected sheet names in a dictionary format like {defaultsheetname=expectedsheetnames}'''
            for row in range(1, noOfRows):
                cellval= worksheet.cell_value(row,0)
                
                defaultValue = worksheet.cell_value(row,1)
                sheetNamedictVar[str(defaultValue)] = str(cellval)

            print 'Sheetname mapping dictionary:\n',sheetNamedictVar,'\n'

            '''verify that the sheet name in client file matches with the expected sheet names in the config file and return the default sheet name if matches or return the same client sheet name if not matches'''
            for sheetName in sheetNamedictVar:
                cellval = sheetNamedictVar[sheetName]
                cellval = cellval.split(",")            
                if sname in cellval:
                    print 'Default sheet name for: ',sname,' is: ',sheetName,'\n'
                    return sheetName,'True','Pass'
            return 'None','False','Client file Sheet name is not matching with any of the expected sheet names in config file. Sheet Name in client file is '+str(sname)
        except Exception, e:
            print e
            return 'None','False','Exception in getting the default sheet name from config file for the Sheet Name in client file is '+str(sname)+' ERROR: '+str(e)
            #raise ValueError("Sheet name is not matching with any of the expected sheet names in config file. Sheet Name in client file ",sname)
    def get_excel_no_of_rows_and_cols(self,file_path,sheet_name='None'):
        try:
            
            workb = xlrd.open_workbook(file_path)
            if sheet_name == 'None':
                worksheet = workb.sheet_by_index(0)
            else:
                worksheet = workb.sheet_by_name(sheet_name)
            noofrows = worksheet.nrows
            noofcols = worksheet.ncols
            print 'Rows, cols:',noofrows,noofcols
            return noofrows,noofcols
        except Exception, e:
            print 'Exception',e
            return 'None','Exception in reading the file: '+file_path+' to get rows and cols. ERROR: '+str(e)


    def write_sheet_data_to_macro_file(self,macro_file_path,sheet_data_to_fill,file_path_to_save,policyFolderPath):
        try:
            ''' reading the file using the xlrd to get the row count as using openpyxl getting wrong row count '''
            workb = xlrd.open_workbook(macro_file_path)
            worksheet=workb.sheet_by_index(0)
            noofrows=worksheet.nrows

            '''Get the excel as worksheet to write the data in it using openpyxl and in the first sheet'''
            wb = load_workbook(macro_file_path, keep_vba=True)
            print(wb.sheetnames)
            sheet= wb.sheetnames[0]
            print sheet
            ws= wb.get_active_sheet()
            '''Get the columns count to get the header values'''
            column_count = ws.max_column
            print column_count

            '''Create a dictioanry for storing the header vcalues with the index'''
            macroHeaderDict={}
            for index in range (1,column_count+1):
                key=ws.cell(row=2, column=index,).value
                macroHeaderDict[str(key)]=int(index)
            print macroHeaderDict

            ''' Fill data in macro sheet row by row'''
            '''sheet_data_to_fill dict vaiable contain the values in a format of {1={col1=value1,col2=val2},2={col1=val1,col3=val3}} writing each internal dict into a row'''
            for data in sheet_data_to_fill:
                '''for every data in sheet data dictionary writing in a new row'''
                noofrows=noofrows+1
                print 'Row Count',noofrows
                required_data = sheet_data_to_fill[data]
                for header in required_data:
                    print 'header ',header
                    try:
                        col=macroHeaderDict[header]
                    except:
                        print 'Header :',header,' is not available in macro file\n'
                        continue
                    value = required_data[header]
                    if header == 'Salary /Earning':
                        value = float(value)
                        value = round(value,2)
                    print 'writing the value ',value,' in ',col,' column index for header: ',header, ' in row: ',noofrows+1
                    ws.cell(row=noofrows, column=col).value = value
            if not os.path.exists(policyFolderPath):
                os.makedirs(policyFolderPath) 
            wb.save(file_path_to_save)
            return 'True','Pass'
        except Exception, e:
            print e
            return 'False','File exception in write sheet data to standard macro file template '+str(macro_file_path)+' Error: '+str(e)
            raise ValueError ('File exception in write sheet data to macro file keyword ',macro_file_path)
        
    def write_error_msg_to_excel_sheet(self,folder_path,file_path, policy_number,sheet_name,process_name,status,err_msg, emp_code = ''):
        try:
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
                
            if not os.path.exists(file_path):
                print 'File Not exists'
                work_b = Workbook()
                work_b.save(file_path)
                noofrows = 1
            else:
                print 'File exists'
                workb = xlrd.open_workbook(file_path)
                worksheet = workb.sheet_by_index(0)
                noofrows = worksheet.nrows
                noofcols = worksheet.ncols
            wb = load_workbook(file_path)
            print(wb.sheetnames)
            sheet= wb.sheetnames[0]
            ws= wb.get_active_sheet()
            if noofrows == 0:
                noofrows = noofrows+2
            else:
                noofrows = noofrows+1
            

            ws.cell(row = 1, column = 1).value = 'Policy Number'
            ws.cell(row = 1, column = 2).value = 'Sheet Name'
            ws.cell(row = 1, column = 3).value = 'Process Name'
            ws.cell(row = 1, column = 4).value = 'Status'
            ws.cell(row = 1, column = 5).value = 'Error Message'
            ws.cell(row = 1, column = 6).value = 'Employee Code or Schedule Number'
            
            
            
            ws.cell(row = noofrows, column = 1).value = policy_number
            ws.cell(row = noofrows, column = 2).value = sheet_name
            ws.cell(row = noofrows, column = 3).value = process_name
            ws.cell(row = noofrows, column = 4).value = status
            ws.cell(row = noofrows, column = 5).value = err_msg
            ws.cell(row = noofrows, column = 6).value = emp_code
            
            wb.save(file_path)
            return 'True','Pass'
        except Exception, e:
            print 'Exception in writig the error file. ERROR: ',e
            return 'False','Exception in writig the error file. ERROR: '+str(e)
            raise ValueError ('Cannot write the data into error file')   
        

    def move_created_files_to_timestamp_folder(self,source_folder_path,destination_folder_Path,time_stamp):
        try:
        
            #time_stamp = time.strftime("%d-%m-%y-%H-%M-%S")
            print 'time stamp: ',time_stamp
            destination_folder_with_time_stamp = destination_folder_Path+'\\'+str(time_stamp)
            print 'destination_folder_with_time_stamp: ',destination_folder_with_time_stamp
            os.makedirs(destination_folder_with_time_stamp)
            print 'destination file path',destination_folder_with_time_stamp
            files = os.listdir(source_folder_path)
            #print len(files)
            if len(files) == 0:
                print 'there are no files in directory'+path
            #files.sort()
            for f in files:
                print f
                src = os.path.join(source_folder_path, f)
                print 'source: ',src
                dst = os.path.join(destination_folder_with_time_stamp, f)
                print 'destination: '+dst
                sheetsmoved = shutil.move(src,dst)
                print sheetsmoved
        except Exception, e:
            print e
            raise ValueError('Fail to move created files to time stam folder',source_folder_path)
        #os.rename("C:\\FG1","C:\\FG2")
    def move_selected_file_to_required_folder(self,src_file,dest_folder,time_stamp='None'):
        try:
            
            if time_stamp == 'None':
                time_stamp = time.strftime("%d-%m-%y-%H-%M-%S")
            print 'time stamp: ',time_stamp
            destination_folder_with_time_stamp = dest_folder+'\\'+str(time_stamp)
            print 'destination_folder_with_time_stamp: ',destination_folder_with_time_stamp
            if not os.path.exists(destination_folder_with_time_stamp):
                os.makedirs(destination_folder_with_time_stamp)
            file_names = src_file.split('\\')
            len_file = len(file_names)
            file_name = file_names[len_file-1]
            if os.path.exists(destination_folder_with_time_stamp+'\\'+file_name):
                os.remove(destination_folder_with_time_stamp+'\\'+file_name)
            
            #src = os.path.join(source_folder_path, src_file)
            #print 'source: ',src
            #dst = os.path.join(destination_folder_with_time_stamp, src_file)
            #print 'destination: '+dst
            sheetsmoved = shutil.move(src_file,destination_folder_with_time_stamp)
            return True,'Pass'
        except Exception, e:
            return False,' Cannot able move the file '+str(src_file)+' to folder '+str(dest_folder)+'. ERROR: '+str(e)

    def get_static_values_into_dictionary(self,config_file_path,defaultSheetName):
        try:
            print ''' Add static values from configuration file staticdata sheet into di ctionary'''
            workbook=xlrd.open_workbook(config_file_path)
            sh=workbook.sheet_by_name('StaticData')
            row_count = sh.nrows
            column_count = sh.ncols
            print row_count
            print column_count
            dictvar={}
            
            for rowindex in range(1,row_count):
                #for colindex in range(0,column_count):
                KeyValues = sh.cell(rowindex,0).value
                keyColumnValues = sh.cell(rowindex,1).value
                print KeyValues
                print keyColumnValues
                dictvar[str(KeyValues)]=str(keyColumnValues)
            print ''' Adding Transaction type based on sheet name into dictoinary as static values'''
            trans_type = 'Tr.\nType'
            print trans_type
            print 'defaultSheetName',defaultSheetName
            if defaultSheetName == 'Additions':
                dictvar[str(trans_type)] = str('A')
            if defaultSheetName == 'Deletion':
                dictvar[str(trans_type)] = 'T'
            if defaultSheetName == 'Transfer' or defaultSheetName == 'Revisions' or defaultSheetName == 'Data Correction':
                dictvar[str(trans_type)] = 'C'
            return dictvar,'True','Pass'
        except Exception, e:
            print 'Exception in adding static values into dictionary Error: ',e
            return 'None','False','Exception in getting the static values from configuration file for the client file sheet: '+str(defaultSheetName)+'. ERROR: '+str(e)
            #raise ValueError('get static values into dictionary from configuration file')
    
    def get_subsidiary_data_into_dictonary(self,macro_file_column_name,column_value_in_client,lookupdata_file_path):
        try:
            print ''' Map Master data with the client data from LookupData file'''
            workbook=xlrd.open_workbook(lookupdata_file_path)
            sh=workbook.sheet_by_name('LookupData')
            row_count = sh.nrows
            column_count = sh.ncols
            print 'Lookup data row count: ',row_count
            print 'Lookup data col count: ',column_count
            dictvar={}
            if (macro_file_column_name == 'Subsidiary'):
                for rowindex in range(1,row_count):
                    KeyValues = sh.cell(rowindex,0).value
                    keyColumnValues = sh.cell(rowindex,1).value
                    print KeyValues
                    print keyColumnValues
                    dictvar[str(KeyValues)]= keyColumnValues
            try:
                
                newvalue = dictvar[str(column_value_in_client)]
                print '''newvalue'''
                print newvalue
                   
                return newvalue
            except Exception, e:
                print e
                return column_value_in_client
        except Exception, e:
            print 'Exception in mapping the master data(Subsidary, etc) to dictionary. ERROR: ',e
            raise ValueError('get subsidiary data into dictionary',lookupdata_file_path)
        '''if colname = 'policy':
            for rowindex in range(0,row_count):
            for colindex in range(0,column_count):
                KeyValues = sh.cell(rowindex,2).value
                keyColumnValues = sh.cell(rowindex,3).value
            print KeyValues
            print keyColumnValues
            dictvar[str(KeyValues)]=str(keyColumnValues)
        try:
            newvalue = dictVar[value]
            return newvalue
        except:
            return column_value'''
        
    def format_date_to_dd_mm_yyyy(self,sheet_name,client_column_name, column_value, macro_col_name,no_of_days_to_add,workbook):
        try:
            
            print 'Format Dates using the client column name'
            print 'column_value',column_value
            #typeof = type(column_value_in_client).__name__
            try:
                print 'Entered in date convertor'
                column_value_in_client = datetime(*xlrd.xldate_as_tuple(column_value, workbook.datemode))
                print 'Converted tuple date: ',column_value_in_client
                column_value = datetime.strftime(column_value_in_client,'%d-%m-%Y')
                status = 'true'
            except Exception,e:
                print 'Error in converting date format to tuple date: ',e
                print 'client_column_name ',client_column_name
                client_column_name = client_column_name.lower()
                status = 'false'
                try:
                    if 'dd/mmm/yyyy' in client_column_name:
                        print 'Entered in DD/MMM/YYYY and column_value: ',column_value
                        column_value = datetime.strptime(column_value,'%d/%b/%Y').strftime('%d-%m-%Y')
                        print 'After entered  in DD/MMM/YYYY and column_value: ',column_value
                        status = 'true'
                    if 'dd/mm/yyyy' in client_column_name:
                        print 'Entered in DD/MM/YYYY and column_value: ',column_value
                        column_value = datetime.strptime(column_value,'%d/%m/%Y').strftime('%d-%m-%Y')
                        print 'After entered  in DD/MM/YYYY and column_value: ',column_value
                        status = 'true'
                    if 'mm/dd/yyyy' in client_column_name:
                        print 'Entered in MM/DD/YYYY and column_value: ',column_value
                        column_value = datetime.strptime(column_value,'%m/%d/%Y').strftime('%d-%m-%Y')
                        print 'After entered  in MM/DD/YYYY and column_value: ',column_value
                        status = 'true'
                    if 'mmm/dd/yyyy' in client_column_name:
                        print 'Entered in MMM/DD/YYYY and column_value: ',column_value
                        column_value = datetime.strptime(column_value,'%b/%d/%Y').strftime('%d-%m-%Y')
                        print 'After entered  in MMM/DD/YYYY and column_value: ',column_value
                        status = 'true'
                    if 'dd-mmm-yyyy' in client_column_name:
                        print 'Entered in DD-MMM-YYYY and column_value: ',column_value
                        column_value = datetime.strptime(column_value,'%d-%b-%Y').strftime('%d-%m-%Y')
                        print 'After entered  in DD-MMM-YYYY and column_value: ',column_value
                        status = 'true'
                    if 'mm-dd-yyyy' in client_column_name:
                        print 'Entered in MM-DD-YYYY and column_value: ',column_value
                        column_value = datetime.strptime(column_value,'%m-%d-%Y').strftime('%d-%m-%Y')
                        print 'After entered  in MM-DD-YYYY and column_value: ',column_value
                        status = 'true'
                    if 'mmm-dd-yyyy' in client_column_name:
                        print 'Entered in MMM-DD-YYYY and column_value: ',column_value
                        column_value = datetime.strptime(column_value,'%b-%d-%Y').strftime('%d-%m-%Y')
                        print 'After entered  in MMM-DD-YYYY and column_value: ',column_value
                        status = 'true'

                except:
                    try:
                        print 'Entered in DD/MMM/YYYY and column_value: ',column_value
                        column_value = datetime.strptime(column_value,'%d/%b/%Y').strftime('%d-%m-%Y')
                        print 'After entered  in DD/MMM/YYYY and column_value: ',column_value
                        status = 'true'
                    except:
                        try:
                            print 'Entered in DD-MMM-YYYY and column_value: ',column_value
                            column_value = datetime.strptime(column_value,'%d-%b-%Y').strftime('%d-%m-%Y')
                            print 'After entered  in DD-MMM-YYYY and column_value: ',column_value
                            status = 'true'
                            
                        except:
                            try:
                                print 'Entered in DD/MM/YYYY and column_value: ',column_value
                                column_value = datetime.strptime(column_value,'%d/%m/%Y').strftime('%d-%m-%Y')
                                print 'After entered  in DD/MM/YYYY and column_value: ',column_value
                                status = 'true'
                            except:
                                try:                                    
                                    print 'Entered in DD-MM-YYYY and column_value: ',column_value
                                    column_value = datetime.strptime(column_value,'%d-%m-%Y').strftime('%d-%m-%Y')
                                    print 'After entered  in DD/MM/YYYY and column_value: ',column_value
                                    status = 'true'
                                except:
                                    try:                                    
                                        print 'Entered in MMM/DD/YYYY and column_value: ',column_value
                                        column_value = datetime.strptime(column_value,'%b/%d/%Y').strftime('%d-%m-%Y')
                                        print 'After entered  in MMM/DD/YYYY and column_value: ',column_value
                                        status = 'true'
                                    except:
                                        try:                                        
                                            print 'Entered in MM/DD/YYYY and column_value: ',column_value
                                            column_value = datetime.strptime(column_value,'%m/%d/%Y').strftime('%d-%m-%Y')
                                            print 'After entered  in MM/DD/YYYY and column_value: ',column_value
                                            status = 'true'
                                        except:
                                            try:                                            
                                                print 'Entered in MM-DD-YYYY and column_value: ',column_value
                                                column_value = datetime.strptime(column_value,'%m-%d-%Y').strftime('%d-%m-%Y')
                                                print 'After entered  in MM-DD-YYYY and column_value: ',column_value
                                                status = 'true'
                                            except:
                                                print 'Entered in MMM-DD-YYYY and column_value: ',column_value
                                                column_value = datetime.strptime(column_value,'%b-%d-%Y').strftime('%d-%m-%Y')
                                                print 'After entered  in MMM-DD-YYYY and column_value: ',column_value
                                                status = 'true'                  
            
            print 'column_value',column_value
            if status == 'false':
                return 'None','False','Dates are not matching any of the formats client Date: '+str(column_value)
                #raise ValueError ('Dates are not matching any of the formats column value',)
                
            if sheet_name == 'Deletion' and macro_col_name == 'Effective\n Date':
                column_value = self.increase_date(column_value,no_of_days_to_add)
            return column_value,'True','Pass'
        except Exception, e:
            print ('Error in formatting Dates ',e)
            return 'None','False','Exception in converting the given date format in client file sheet_name: '+sheet_name+',client_column_name: '+client_column_name+', column_value: '+str(column_value)+'. ERROR: '+str(e)
            #raise ValueError ('Error in formatting Dates ',e)
            
    def increase_date(self,cdate,noofdays):
        try:
            #cdate = datetime.now()
            cdate = datetime.strptime(cdate,'%d-%m-%Y')
            frpdate = cdate + timedelta(int(noofdays))
            frpdate = frpdate.strftime("%d-%m-%Y")
            print frpdate
            return frpdate
        except Exception, e:
            print 'Exception in increasing date by ',noofdays,'Error: ',e
            return 'None','Exception in increase/decrease the date: '+str(cdate)+'. ERROR: '+str(e)
            #raise ValueError ('Exception in increasing date by ',noofdays,'Error: ',e)

    def get_ms_excel_multiple_row_values_into_dictionary_list_based_on_key(self,filepath,keyName,sheetName=None):
            """Returns the dictionary of values given row in the MS Excel file """
            try:
                
                workbook = xlrd.open_workbook(filepath)
                snames=workbook.sheet_names()
                if sheetName==None:
                    sheetName=snames[0]      
                worksheet=workbook.sheet_by_name(sheetName)
                noofrows=worksheet.nrows
                listDict = {}
                size = int(0);
                headersList=worksheet.row_values(int(0))
                for rowNo in range(1,int(noofrows)):
                    dictVar={}
                    rowValues=worksheet.row_values(int(rowNo))                
                    if str(rowValues[0])!=str(keyName):
                        continue
                    size = size+1		
                    for rowIndex in range(0,len(rowValues)):
                        cell_data=rowValues[rowIndex]               
                        dictVar[str(headersList[rowIndex])]=cell_data
                    listDict[size] = dictVar
                return listDict,'True','Pass'
            except:
                print 'Exception in getting excel multiple row values into dictionary for the key: '+str(keyName)
                return 'None','False','Exception in getting excel file '+filepath+'  row values into dictionary for the key: '+str(keyName)+' ERROR: '+str(e)
    def get_ms_excel_row_values_into_dictionary_based_on_key(self,filepath,keyName,sheetName=None):
            """Returns the dictionary of values given row in the MS Excel file """
            try:
                
                workbook = xlrd.open_workbook(filepath)
                snames=workbook.sheet_names()
                dictVar={}
                if sheetName==None:
                    sheetName=snames[0]      
                worksheet=workbook.sheet_by_name(sheetName)
                noofrows=worksheet.nrows
                dictVar={}
                headersList=worksheet.row_values(int(0))
                for rowNo in range(1,int(noofrows)):
                    rowValues=worksheet.row_values(int(rowNo))
                    if str(rowValues[0])==str(keyName):
                        for colIndex in range(0,len(rowValues)):
                            cell_data=rowValues[colIndex]                
                            dictVar[str(headersList[colIndex])]=str(cell_data)
                        continue
                    
                return dictVar,'True','Pass'
            except Exception, e:
                print 'Exception in getting the row values into dictionary based on key: '+str(keyName)
                return 'None','False','Exception in getting the row values into dictionary based on key: '+str(keyName)+' ERROR: '+str(e)
    
    def mapping_master_file_data_for_plan_and_occp(self,mul_dict,policy_number,subsidary_name,cover_base,grade_status,subsidiary_value_status,structure_name=''):
        try:
            
            dictVar = {}
            print 'mul_dict\n',mul_dict,'policy_number\n',policy_number,'subsidary_name\n',subsidary_name,'structure_name\n',structure_name,',cover_base,grade_status',cover_base,grade_status
            subsidary_status = 'False'
            struc_status = 'False'
            #print 'multiple dictionary',mul_dict
            for key in mul_dict:
                ind_dict = mul_dict[key]
                print 'ind_dict',ind_dict
                if ind_dict['Master Policy'] == policy_number:
                    if ind_dict['Susbsidiary Name'].replace(" ", "").lower() == subsidary_name.replace(" ", "").lower():
                        subsidary_status = 'True'
                        dictVar['City'] = ind_dict['City']
                        dictVar['Z'] = int(ind_dict['Z'])
                        dictVar['Subsidiary'] = int(ind_dict['Susbsidiary No'])
                        if int(cover_base) == 3 and grade_status == 'True':
                            print 'Struture type$$$$$$$$$$$$$$',ind_dict['Structure']
                            band_type = ind_dict['Structure'].split(",")
                            print 'band_type@@@@@@@@@@@@',band_type
                            for stru_type in band_type:
                               print'stru_type^^^^^^^^^^^^^^^^^',stru_type
                               if structure_name.replace(" ", "").lower() == stru_type.replace(" ","").lower():
                                   
                                   print 'stru_type!!!!!!!!!!!!!!!!!!!!!!!!!!',stru_type
                                   print'Structure***********************',structure_name.replace(" ", "").lower() == stru_type.replace(" ","").lower()
                                   print '\n++++++++++++++++Entered in structure++++++++++++++++++++++=\n'
                                   struc_status = 'True'
                                   dictVar['Plan'] = ind_dict['Plan']
                                   dictVar['OCCp'] = ind_dict['OCCp']
                                   print 'Plan Occp Structure :',dictVar
                                   return dictVar,'True','Pass'
                                   break
                        else:
                            dictVar['Plan'] = ind_dict['Plan']
                            dictVar['OCCp'] = ind_dict['OCCp']
                            return dictVar,'True','Pass'
            if int(cover_base) == 3 and grade_status == 'True' and struc_status == 'False':
                err_msg = 'Structure/Band name \''+structure_name+'\' is not matching in the Master data file for policy number: '+policy_number
                return 'None','False',err_msg
            if subsidiary_value_status == 'True' and subsidary_status == 'False':
                return 'None','False','Subsidiary Name: '+str(subsidary_name)+' is not matching in the Master data file for the policy'
            if subsidiary_value_status == 'False' and subsidary_status == 'False':
                if len(mul_dict) == 1:
                    ind_dict = mul_dict[1]
                    dictVar['City'] = ind_dict['City']
                    dictVar['Z'] = int(ind_dict['Z'])
                    dictVar['Plan'] = ind_dict['Plan']
                    dictVar['OCCp'] = ind_dict['OCCp']
                    dictVar['Subsidiary'] = int(ind_dict['Susbsidiary No'])
                    print 'City Z Plan Occp Subsidiary :',dictVar
                    return dictVar,'True','Pass'
                else:
                    return dictVar, 'False', 'Multiple rows of data available for the policy number in Master data file and subsidiary name not provided in client data'
            return 'None','False','Mapping of master data failed'
        except Exception , e:
            print 'Exception in mapping the master file data with the client data for policy_number '+policy_number+',subsidary_name '+subsidary_name+',structure_name '+structure_name+' ERROR: '+str(e)
            return 'None','False','Exception in mapping the master file data with the client data for policy_number '+policy_number+',subsidary_name '+subsidary_name+',structure_name '+structure_name+' ERROR: '+str(e)
            
    def get_list_of_directories_in_path(self, directory_path):
        try:
            
            #directory_path = 'C:\\Projects\\FG\\Project Code\\FGE\\TestData\\output_folder'
            list_of_dirs = []
            print os.listdir(directory_path)
            for name in os.listdir(directory_path):
                if os.path.isdir(directory_path+'\\'+name):
                    list_of_dirs.append(name)
            print 'list_dir\n',list_of_dirs
            if len(list_of_dirs) == 0:
                return 'None','Expected directories with policy numbers in the folder '+directory_path+'. Zero policy folders are available'
            return list_of_dirs,'Pass'
        except Exception, e:
            return 'None','Exception in getting the directories of created policy macro files. ERROR: '+str(e)
        
    def get_list_of_macro_files_in_path(self,directory_path,sub_directory_name):
        try:
            directory_path = directory_path+'\\'+sub_directory_name
            #directory_path = 'C:\\Projects\\FG\\Project Code\\FGE\\TestData\\output_folder\\GL000460'
            list_of_files = []
            extension = '.xlsm'
            print os.listdir(directory_path)
            for name in os.listdir(directory_path):
                if name.endswith(extension):
                    list_of_files.append(name)
            print 'list_dir\n',list_of_files
            if len(list_of_files) == 0:
                return 'None','No files available with extension: '+extension+' in directory: '+directory_path
            file_name_dict = {}
            for f in list_of_files:
                if 'Addition' in f:
                    file_name_dict[2] = directory_path+'\\'+f
                if 'Deletion' in f:
                    file_name_dict[1] = directory_path+'\\'+f
                if 'Revision' in f:
                    file_name_dict[3] = directory_path+'\\'+f
            print 'file_name_dict\n',file_name_dict
            if len(file_name_dict) == 0:
                return 'None','No files available with names Addition, Deletion and Revision in folder: '+directory_path+' available files are: '+list_of_files
            
            return file_name_dict,directory_path,'Pass'
        except Exception, e :
            return 'None','None','Exception in getting the macro file list from the: '+str(directory_path)+' folder for policy '+str(sub_directory_name)+' . ERROR: '+str(e)
        
    def  validate_generated_bill_reports(self,uploadedmacrofilepath,extractedbillfilepath,columnName='None'):
        try:
            xl_file_list1 = []
            xl_file_list2 = []
            ''' reading the file using the xlrd to get the row count as using openpyxl getting wrong row count for uploadedmacrofile '''
            workbook1 = xlrd.open_workbook(uploadedmacrofilepath)
            worksheet1 = workbook1.sheet_by_name('Sheet1')
            noofrows1=worksheet1.nrows
            noofrows= noofrows1-6   
            noofcols1=worksheet1.ncols
            print 'Entered'
            ''' reading the file using the xlrd to get the row count as using openpyxl getting wrong row count for extractedbillfile '''
            workbook2 = xlrd.open_workbook(extractedbillfilepath)
            #workbook2 = os.path.join(extractedbillfilepath)
            worksheet2 = workbook2.sheet_by_name('11384')
            noofrows2=worksheet2.nrows
            noofrows2=noofrows2-1
            
            noofcols2=worksheet2.ncols
            ''' comparing number of rows in uploadedmacrofile and extractedbillfile'''
            print 'comparing number of rows in uploadedmacrofile and extractedbillfile'
            if noofrows != noofrows2:
                print 'excel files are miss match :'
                return False, 'No of rows from macro excel:'+str(noofrows)+' are not matching with the Bill report excel rows: '+str(noofrows2)
                
            ''' getting  the header names and column values into for uploadedmacrofile '''
            headersList1,status, msg = self.get_client_file_header_values_into_dict(worksheet1,1,'Sheet1')
            print 'headersList1',headersList1
            columnName = 'Emp\nNo'
            colIndex1 = headersList1[columnName]
            #print 'columnName: ',colIndex2
            print 'colIndex1 ',colIndex1
            for rowNo in range(6,int(noofrows1)):
                rowValue1=worksheet1.cell_value(rowNo,colIndex1)
                
                xl_file_list1.append(str(rowValue1.strip()))
                print 'append rowValue1 ',rowValue1
            print 'List values : ',xl_file_list1
            ''' To get columnvalues in excel sheet2'''

            ''' getting  the header names and column values into for extractedbillfile ''' 
            #columnName=str(columnName)
            headersList,status, msg  = self.get_client_file_header_values_into_dict(worksheet2,0,'Sheet1')
            #print 'headersList 2',headersList
            col_name_bill = 'EMPNO'
            colIndex2 = headersList[col_name_bill]
            for rowNo in range(1,int(noofrows2)+1):
                rowValue2=worksheet2.cell_value(rowNo,colIndex2)
                
                xl_file_list2.append(str(rowValue2.strip()))
            print xl_file_list2
            if len(xl_file_list1) != len(xl_file_list2):
                return False,'Number of employee code from the macro file:'+str(xl_file_list1)+' is not matching with the Employee Codes in bill report'+str(xl_file_list2)
            for empno_List in xl_file_list1:
                print 'emp list value :\n',empno_List
                if empno_List in xl_file_list2:
                    print 'list2:\n',empno_List
                    continue
                else:
                    return  False, 'Employee code from macro file: '+str(empno_List)+' is not in Bill report employee code list '+str(xl_file_list2)
            return True,'Pass'
        except Exception, e:
            return False, 'Exception in validating the Bill Report Employee Code with the macro file employee code. ERROR: '+str(e)
    def get_subsidary_values(self,marco_file_path,marco_sheet_name='None'):
        try:
                    
            workbook = xlrd.open_workbook(marco_file_path)
            if marco_sheet_name == 'None':
                sheet = workbook.sheet_by_index(0)
            else:
                sheet = workbook.sheet_by_name(marco_sheet_name)
            
            subsidarylist=[]
            x = {}
            sublist =[]
            for rownum in range(6,sheet.nrows):
                rowValues=sheet.row_values(int(rownum))
                x[rowValues[3]]=rownum
            print(x)

            for key in x:
                print(x[key])
                #print 'dic value : ', key
                sublist.append(key)
                print sublist
            return sublist,'Pass'
        except Exception, e:
            return 'None','Exception in getting the subsidary names from the macro excel file. ERROR: '+str(e)
    def get_ms_excel_multiple_row_values_into_dictionary_list_based_on_key_in_any_col(self,extractedbillnumberpath,keyName,sheetName=None):
            """Returns the dictionary of values given row in the MS Excel file """
            try:
                workbook = xlrd.open_workbook(extractedbillnumberpath)
                snames=workbook.sheet_names()
                if sheetName==None:
                    sheetName=snames[0]      
                worksheet=workbook.sheet_by_name(sheetName)
                noofrows=worksheet.nrows
                listDict = {}
                size = int(0);
                headersList=worksheet.row_values(int(0))
                for rowNo in range(1,int(noofrows)):
                    dictVar={}
                    rowValues=worksheet.row_values(int(rowNo))
                    status = 'false'
                    for row in rowValues:
                        if str(row)!=str(keyName):
                            continue
                        else:
                            status = 'true'
                            break
                    if status == 'false':
                        continue
                    
                    size = size+1  
                    for rowIndex in range(0,len(rowValues)):
                        cell_data=rowValues[rowIndex]               
                        dictVar[str(headersList[rowIndex])]=cell_data
                    listDict[size] = dictVar
                return listDict,'True','Pass'
            except Exception, e:
                print 'Exception in getting excel multiple row values into dictionary for the billnumber: '+str(keyName)
                return 'None','False','Exception in getting excel file '+extractedbillnumberpath,+'  row values into dictionary for the key: '+str(keyName)+' ERROR: '+str(e)
    def get_bill_number_based_on_time_stamp(self,file_path, policy_num,sheet_name='None'):
        try:
                
            bill_num_lsit = []
            workbook = xlrd.open_workbook(file_path)
            if sheet_name == 'None':
                sheet = workbook.sheet_by_index(0)
            else:
                sheet = workbook.sheet_by_name(sheet_name)
            no_of_rows = sheet.nrows
            headersList,status, msg  = self.get_client_file_header_values_into_dict(sheet,0,'Sheet1')
            if status == 'False':
                return 'None','False',msg
            col_index = headersList['DATIME']
            print 'no_of_rows: ',no_of_rows,'col_index: ',col_index
            datetime_key = sheet.cell_value(no_of_rows-1, col_index)
            mul_dict_of_same_time_stamp, status, msg = self.get_ms_excel_multiple_row_values_into_dictionary_list_based_on_key_in_any_col(file_path,datetime_key)
            if status == 'False':
                return 'None','False',msg
            print 'mul_dict_of_same_time_stamp:\n',mul_dict_of_same_time_stamp
            for index in mul_dict_of_same_time_stamp:
                ind_dict = mul_dict_of_same_time_stamp[index]
                p_num = ind_dict['CHRDNUM']
                if str(p_num) != str(policy_num):
                    return 'None','False','Transaction Policy number: '+policy_num+' is not matching with DTF billnumber file last DATIME: '+datetime_key+' policy number: '+p_num+'. Or Policy numbers are different for same time stamp'
                bill_num = ind_dict['BILLNO']
                bill_num_lsit.append(int(bill_num))
            print 'bill_num_lsit:\n',bill_num_lsit
            return bill_num_lsit, 'True', 'Pass'
        except Exception, e:
            return 'None','Exception in getting the bill number based on the stamp. ERROR: '+str(e)
        
    def verify_and_write_error_records_to_log(self,status, error_record_list,err_msg,folder_path,file_path, policy_number,sheet_name,schedule_num):
        try:
                
            if status == 'Fail':
                if error_record_list != 'None':
                    for record in error_record_list:
                        self.write_error_msg_to_excel_sheet(folder_path,file_path, policy_number,sheet_name,'Spool File Verification','False',str(record),schedule_num)
                
                self.write_error_msg_to_excel_sheet(folder_path,file_path, policy_number,sheet_name,'Spool File Verification','False',err_msg,schedule_num)
                return 'Fail','verification in spool file failed. '+str(err_msg)
            if status == 'False':
                if error_record_list != 'None':
                    for record in error_record_list:
                        self.write_error_msg_to_excel_sheet(folder_path,file_path, policy_number,sheet_name,'Spool File Verification','False',str(record),schedule_num)
                
                self.write_error_msg_to_excel_sheet(folder_path,file_path, policy_number,sheet_name,'Spool File Verification','False',err_msg,schedule_num)
                return 'True','Pass'
            if status == 'True':
                return 'True','Pass'
        except Exception, e:
            return 'False','Exception in Writing the Error records to Failed Logs. ERROR: '+str(e)
        
            
    def move_folders(self,src,dst,time_stamp='None'):
        try:
            if time_stamp == 'None':
                time_stamp = time.strftime("%d-%m-%y-%H-%M-%S")
            destination_folder_with_time_stamp = dst+'\\'+str(time_stamp)
            print 'destination_folder_with_time_stamp: ',destination_folder_with_time_stamp
            if not os.path.exists(destination_folder_with_time_stamp):
                os.makedirs(destination_folder_with_time_stamp)
            for node in os.listdir(src):
                print 'node ',node
                shutil.move(os.path.join(src, node) , os.path.join(destination_folder_with_time_stamp, node))
            return 'True','Pass'
        except Exception, e:
            print 'Exception ',str(e)
            return 'False','Exception in moving the macro file fodlers to destination path: '+str(dst)+' with time stamp'
    def get_ms_excel_multiple_row_values_into_dictionary_list_based_on_key_for_billnumber(self,extractedbillnumberpath,sheetName=None):
            """Returns the dictionary of values given row in the MS Excel file """
            try:           
                workbook = xlrd.open_workbook(extractedbillnumberpath)
                snames=workbook.sheet_names()
                if sheetName==None:
                    sheetName=snames[0]
                    print 'sheetname',sheetName
                worksheet=workbook.sheet_by_name(sheetName)
                noofrows=worksheet.nrows
                print 'noofrows',noofrows
                listDict = {}
                size = int(0);
                print 'size',size
                headersList=worksheet.row_values(int(0))
                print 'hearderList',headersList
                for rowNo in range(1,int(noofrows)):
                    dictVar={}
                    rowValues=worksheet.row_values(int(rowNo))
                    print 'rowValues',rowValues
                    if ('Yes' in rowValues):
                     rowValues='Do not Execute'  
                     
                   
                    size = size+1
                    print 'size1',size
                    for rowIndex in range(0,len(rowValues)):
                        cell_data=rowValues[rowIndex]
                        print 'cell_data',cell_data
                        dictVar[str(headersList[rowIndex])]=cell_data
                    listDict[size] = dictVar
                print 'listDict',listDict
                return listDict,'True','Pass'
            except Exception, e:
                print 'Exception in getting excel multiple row values into dictionary for the billnumber: '+extractedbillnumberpath
                return 'None','False','Exception in getting excel multiple row values into dictionary for the billnumber '+extractedbillnumberpath+ '. ERROR: '+str(e)
    def get_billnumber_dict(self,listDict):
        try:            
            bill_list = []
            Policy_number = []
            billDict = {}
            '''print 'config_noOfRows: ',config_noOfRows'''
            for number in listDict:
                inner_dict = listDict[number]
                print 'innerdict: ',inner_dict
                #submember = inner_dict['SUBSNUM']
                #print 'submember: ',submember
                #if submember in sub_list:
                bill = inner_dict['BILLNO']
                print 'bill: ',bill
                bill_list.append(bill)
                policy = inner_dict['CHDRNUM']
                print 'policy: ',policy
                
                if 'GL' in policy:
                    Policy_number.append(policy)
                    print Policy_number
                    print bill_list
                #dictVar[str(macro_file_column_name)] = str(column_value_in_client)
                    billDict[int(bill)] = str(policy)
                    print 'billdict: ',billDict
                else:
                    print 'Policy number contains other than GL :',policy
            return billDict
            #return billDict,bill_list,Policy_number
        except Exception, e:
            print 'Exception in getting the billnumber dict: '+str(listDict)+'with dictionary contains no billnumber,policy number'+str(billDict)
            return 'None','False','Exception in getting the billnumber_dict. ERROR: '+str(e)

    def create_a_folder_with_current_date(self,timestampfloderpath):

        time_stamp = time.strftime("%d-%m-%y")
        #os.makedirs(timestampfloderpath+'\\'+str(time_stamp))
        destination_folder_with_time_stamp = timestampfloderpath+'\\'+time_stamp
        print 'destination_folder_with_time_stamp: ',destination_folder_with_time_stamp
        os.makedirs(destination_folder_with_time_stamp)
        return destination_folder_with_time_stamp
        

    def bill_extrct_with_transaction_name(self,filepath):
        try:
            workbook = xlrd.open_workbook(filepath)
            snames=workbook.sheet_names()
            sheetName=snames[0]
            sheetName=sheetName.strip()
            worksheet=workbook.sheet_by_name(sheetName)
            print worksheet
            noofrows=worksheet.nrows
            header_names_dict,status,msg=self.get_client_file_header_values_into_dict(worksheet,0,sheetName)
            print 'header_names_dict/n',header_names_dict
            columnName = 'DESCR'
            colIndex = header_names_dict[columnName]
            print colIndex

            rowValue = worksheet.cell_value(noofrows-1,colIndex)

            if('Terminate Member' in rowValue):
                rowValue = 'Deletion'
                print rowValue
            if('Add Member' in rowValue):
                rowValue = 'Addition'
                print rowValue
            if('Change'in rowValue):
                rowValue = 'Revision'
                print rowValue
            updated_filepath=filepath.replace('.xls',"_")
            
            updated_name= updated_filepath+str(rowValue)+'.xls'
            print updated_name
            print filepath
                    
            os.rename(filepath,updated_name)

        except Exception, e:
            print 'Exception in getting the bill extract with transaction name:. Error: '+str(e)
            return 'None','Exception is getting in bill extraction. Error: '+str(e)

    def write_policy_status_to_excel_sheet(self,folder_path,policy_number,sheet_name,status,err_msg,date_time):
        
        try:
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
            time_stamp = time.strftime("%d-%m-%y")    
            file_path = folder_path+'//'+str(time_stamp)+'.xlsx'   
            if not os.path.exists(file_path):
                print 'File Not exists'
                work_b = Workbook()
                work_b.save(file_path)
                noofrows = 1
            else:
                print 'File exists'
                workb = xlrd.open_workbook(file_path)
                worksheet = workb.sheet_by_index(0)
                noofrows = worksheet.nrows
                noofcols = worksheet.ncols
            wb = load_workbook(file_path)
            print(wb.sheetnames)
            sheet= wb.sheetnames[0]
            ws= wb.get_active_sheet()
            if noofrows == 0:
                noofrows = noofrows+2
            else:
                noofrows = noofrows+1
            

            ws.cell(row = 1, column = 1).value = 'Policy Number'
            ws.cell(row = 1, column = 2).value = 'Sheet Name'
            #ws.cell(row = 1, column = 3).value = 'Process Name'
            ws.cell(row = 1, column = 3).value = 'Status'
            ws.cell(row = 1, column = 4).value = 'Policy Status'
            ws.cell(row = 1, column = 5).value = 'Date Time'
            
            
            
            ws.cell(row = noofrows, column = 1).value = policy_number
            ws.cell(row = noofrows, column = 2).value = sheet_name
            #ws.cell(row = noofrows, column = 3).value = process_name
            ws.cell(row = noofrows, column = 3).value = status
            ws.cell(row = noofrows, column = 4).value = err_msg
            ws.cell(row = noofrows, column = 5).value = date_time
            
            wb.save(file_path)
            return 'True','Pass'
        except Exception, e:
            print 'Exception in writig the error file. ERROR: ',str(e)
            return 'False','Exception in writig the error file. ERROR: '+str(e)
            raise ValueError ('Cannot write the data into error file')
        
    
    def write_error_msg_to_excel_billdata(self,folder_path,file_path,process_name,status,err_msg):
        try:
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
                
            if not os.path.exists(file_path):
                print 'File Not exists'
                work_b = Workbook()
                work_b.save(file_path)
                noofrows = 1
            else:
                print 'File exists'
                workb = xlrd.open_workbook(file_path)
                worksheet = workb.sheet_by_index(0)
                noofrows = worksheet.nrows
                noofcols = worksheet.ncols
            wb = load_workbook(file_path)
            print(wb.sheetnames)
            sheet= wb.sheetnames[0]
            ws= wb.get_active_sheet()
            if noofrows == 0:
                noofrows = noofrows+2
            else:
                noofrows = noofrows+1
            
            ws.cell(row = 1, column = 1).value = 'Process Name'
            ws.cell(row = 1, column = 2).value = 'Status'
            ws.cell(row = 1, column = 3).value = 'Error Message'
          
            ws.cell(row = noofrows, column = 1).value = process_name
            ws.cell(row = noofrows, column = 2).value = status
            ws.cell(row = noofrows, column = 3).value = err_msg
           
            wb.save(file_path)
            return 'True','Pass'
        except Exception, e:
            print 'Exception in writig the error file. ERROR: ',e
            return 'False','Exception in writig the error file. ERROR: '+str(e)
            raise ValueError ('Cannot write the data into error file')   
        
##a = RobotExample()
##extractedbillnumberpath = 'C:\\Users\\anilj\\Desktop\\Billdata.xls'
##extractedbillnumberpath = 'C:\\Users\\anilj\\Desktop\\HDFC\\Sample Data.xls'
##listDict,status,status = a.get_ms_excel_multiple_row_values_into_dictionary_list_based_on_key_for_billnumber(extractedbillnumberpath,sheetName=None)
##billDict = a.get_billnumber_dict(listDict)
##folder_path='C:\\Projects\\FG\\TestData\\policyStatus'
##policy_number='GL000461'
##sheet_name='sheet1'
##status='True'
##err_msg='jhdsgj'
##
##a.write_policy_status_to_excel_sheet(folder_path,policy_number,sheet_name,status,err_msg)

##lookupdata_file_path = 'C:\\Projects\\FG\\TestData\\Master DataFile.xlsx'
##policyNumber = 'GL000463'
##master_file_dict, status, err_msg = a.get_ms_excel_multiple_row_values_into_dictionary_list_based_on_key(lookupdata_file_path,policyNumber)
##print 'master_file_dict, status, err_msg',master_file_dict, status, err_msg
##subsidary_name = 'Planit Testing'
##cover_base = 3
##grade_status = 'True'
##structure_name = 'Graded 3'
##dictV, status,err_msg =a.mapping_master_file_data_for_plan_and_occp(master_file_dict,policyNumber,subsidary_name,cover_base,grade_status,structure_name)
##print 'dictV, status,err_msg',dictV, status,err_msg
##src = 'C:\\FGUpdated\\FG\\TestData\\output_folder\\.'
##dst = 'C:\\FGUpdated\\FG\\TestData\\Moved_macro_folders'
##a.move_folders(src,dst)
#file_path = 'C:\\Users\\gaffurs\\Desktop\\New Microsoft Excel Worksheet.xlsx'
#lis,status,msg = a.get_bill_number_based_on_time_stamp(file_path,'GL00247')
#print 'lis,status,msg: ',lis,status,msg
##uploadedmacrofilepath = 'C:\\Projects\\FG\\FG\\GL000463 Rev.xlsm'
##extractedbillfilepath = 'C:\\Projects\\FG\\FG\\11384.xls'
##status, msg = a.get_subsidary_values(uploadedmacrofilepath, extractedbillfilepath)
##print 'status, msg :',status, msg
#a.get_excel_no_of_rows_and_cols('C:\Projects\FG\FG\TestData\output_folder\GL000460\GL000460_Upload(Additions)25-02-2017.xlsm')
#a.sample()
###a.get_xl_values('C:\Projects\Copy of MemberUploadFile12354.xlsm','Sheet1')
###a.fg_script("C:\Projects\FG\SampleClientData.xls")
###a.fg_script("C:\Projects\Copy of MemberUploadFile12354.xlsm")
###defaultSheetValue = a.get_default_sheet_name_from_configuration_sheet('Additions')
###print 'defaultSheetValue',defaultSheetValue
###a.move_created_files_to_timestamp_folder()
##deafult_client = 'C:\Projects\FG\Project Code\FGE\TestData\client_data'
##config_file = 'C:\Projects\FG\Project Code\FGE\TestData\ConfigurationSheet.xlsx'
##output_folder_to_save_macro = 'C:\Projects\FG\Project Code\FGE\TestData\output_folder'
##macro_file = 'C:\Projects\FG\Project Code\FGE\TestData\Standard_macro.xlsm'
#lookupdata_file_path= 'C:\\Projects\\FG\\Project Code\\FGE\\TestData\\LookupData.xlsx'
##master_data_file = 'C:\\Projects\\FG\\Project Code\\FGE\\TestData\\Master DataFile.xlsx'
##outpu_folder_for_log_errors = 'C:\\Projects\\FG\\Project Code\\FGE\\TestData\\Failed_Consolidation_Logs'
##output_folder_to_move_error_files = 'C:\\Projects\\FG\\Project Code\\FGE\\TestData\\Failed_Consolidation_Files'
##output_folder_for_successfully_consolidated_fies = 'C:\\Projects\\FG\\Project Code\\FGE\\TestData\\Successfuly_Consolidated_files'
#dictVa = a.get_ms_excel_row_values_into_dictionary_based_on_key(master_data_file,'GL000460')
#print dictVa['Cover Basis']

##ict_var = a.mapping_master_file_data_for_plan_and_occp(mul_dict,'GL000463','Planit Team')
##
##print '\n============================================================================\n'
##print dict_var
#a.move_selected_file_to_required_folder('C:\\Projects\\FG\\Project Code\\FGE\\TestData\\client_data\\GL000460.xls','C:\\Projects\\FG\\Project Code\\FGE\\TestData\\Error_files_folder')
#a.read_client_data_files_and_write_to_macro_files_from_default_folder(deafult_client,config_file,output_folder_to_save_macro,macro_file,master_data_file,outpu_folder_for_log_errors,output_folder_to_move_error_files,output_folder_for_successfully_consolidated_fies)
###temp_for_client_files = 'C:\\Users\\gousyas\\Desktop\\FGIFiles\\temp_folder_for_client_files'
##a.move_created_files_to_timestamp_folder(deafult_client,temp_for_client_files)

#temp_for_output_files = 'C:\\Users\\gousyas\\Desktop\\FGIFiles\\temp_for_output_files'
##a.move_created_files_to_timestamp_folder(output_folder,temp_for_output_files)
##column_name = 'DOB (DD/MM/YYYY)'
##col_value = '31/12/2017'
##col_val = a.format_date_to_dd_mm_yyyy('Deletion',column_name,col_value,'Effective\n Date',1)
##print 'updated col value',col_val
##policyNumber = 'GL00460'
##err_dict = {}
##err_dict['Policy Number'] =  policyNumber
##err_dict['Sheet Name'] = 'Addition'
##err_dict['Status'] = 'False'
##err_dict['Error Message'] = 'Failed Sample'
##
##LOG_FILENAME = outpu_folder_for_error_files+'\\ErroFile_'+policyNumber+'.xlsx'
#status , msg = a.write_error_msg_to_excel_sheet(LOG_FILENAME,err_dict)
#print 'status , msg',status , msg


#a.sample()
        
        
