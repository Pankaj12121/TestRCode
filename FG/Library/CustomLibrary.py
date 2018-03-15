import xlrd
import time
import os
from datetime import datetime,timedelta
from win32com.client import Dispatch
from xlwt import Workbook
import win32com.client
from Tkinter import *
from tkMessageBox import *
from openpyxl import load_workbook

class CustomLibrary:

    def __init__(self):
        pass
    def create_result_log_file(self,outputfileName,ResultsDict):
        """Create Results file"""
        try:
            book = Workbook()
            opworksheet = book.add_sheet("Results")
            opworksheet.write(0,0,"BatchName")
            opworksheet.write(0,1,"BatchStatus")
            opworksheet.write(0,2,"Details")
            for rowno in range(1,len(ResultsDict)+1):
                Results = ResultsDict[str(rowno)]
                print Results
                opworksheet.write(rowno,0,Results['BatchName'].replace("$",", "))
                print "",Results['BatchName']
                opworksheet.write(rowno,1,str(Results['BatchStatus']))
                opworksheet.write(rowno,2,Results['BatchDetails'])
            book.save(outputfileName)
        except Exception as exp:
            print "Got exception in read_excel_column_values keyword.Error: "+str(exp)
            raise ValueError('Fail due to no files list available in directiory')
   
            

    def get_time_format_in_script(self,value):
        getNow = datetime.now()
        if "TodayWithoutSlash#" in value:
            subValue = value.split("#")[1]
            dt = datetime(getNow.year,getNow.month,getNow.day)-timedelta(days=int(subValue))
            tempDate = str(dt.year)+str(dt.month)+str(dt.day)
            return tempDate
        if "TodayWithoutSlash$" in value:
            addValue = value.split("$")[1]
            dt = datetime(getNow.year,getNow.month,getNow.day)+timedelta(days=int(addValue))
            tempDate = str(dt.year)+str(dt.month)+str(dt.day)
            return tempDate
        if "TodayWithoutSlash" in value:
            dt = datetime(getNow.year,getNow.month,getNow.day)
            tempDate = str(dt.year)+str(dt.month)+str(dt.day)
            return tempDate
        if "Today#" in value:
            subValue = value.split("#")[1]
            dt = datetime(getNow.year,getNow.month,getNow.day)-timedelta(days=int(subValue))
            tempDate = str(dt.day)+"/"+str(dt.month)+"/"+str(dt.year)
            return tempDate
        if "Today$" in value:
            addValue = value.split("$")[1]
            dt = datetime(getNow.year,getNow.month,getNow.day)+timedelta(days=int(addValue))
            tempDate = str(dt.day)+"/"+str(dt.month)+"/"+str(dt.year)
            return tempDate
        if "Today" in value:
            dt = datetime(getNow.year,getNow.month,getNow.day)
            tempDate = str(dt.day)+"/"+str(dt.month)+"/"+str(dt.year)
            return tempDate
        else:
            return value
            
        
    def read_excel_column_values(self,filepath,sheetname):
        """reads the column values"""
        try:
            workbook = xlrd.open_workbook(filepath)
            worksheet = workbook.sheet_by_name(sheetname)
            noofrows = worksheet.nrows
            colNames=[]
            for rowno in range(0,noofrows):
                cellValue = worksheet.cell_value(rowno,0)
                cellValue = str(cellValue)
                colNames.append(cellValue)
            return colNames
        except Exception as exp:
            print "Got exception in read_excel_column_values keyword.Error: "+str(exp)
            return []
            
                    
        

    def get_latest_file_in_folder(self,folderpath,filestartname='None'):
        fileslist = os.listdir(folderpath)
        screenshotNum = 0
        for fileName in fileslist:
            bStatus = int(str(fileName).find(filestartname))>=0
            if not bStatus:
                continue
            fileNumber = int(str(fileName).replace(".jpg","").split("_")[1])
            if fileNumber > screenshotNum:
                screenshotNum = fileNumber
        filepath = folderpath+"\\screenshot_"+str(screenshotNum)+".jpg"
        print filepath
        if not os.path.exists(filepath):
            return "NA"
        return filepath

    def create_ms_excel_file_using_existing_file(self,inputFilePath,outputFilePath):
        """ It retuen the list of registration codes"""
        book = Workbook()
        workbook = xlrd.open_workbook(inputFilePath)
        snames=workbook.sheet_names()
        expectedColumNumber=-1
        for oldSheetName in snames:
            opworksheet = book.add_sheet(oldSheetName)
            worksheet=workbook.sheet_by_name(oldSheetName)
            noofrows=worksheet.nrows
            tempList=[]
            
            for rowno in range(0,noofrows):
                row=worksheet.row(rowno)
                for colno in range(0,len(row)):
                    cellval=worksheet.cell_value(rowno,colno)
                    if cellval.lower()=='status':
                        expectedColumNumber = colno
                    if colno==expectedColumNumber and rowno >= 1:
                        opworksheet.write(rowno,colno,"Not Executed")
                    else:
                        opworksheet.write(rowno,colno,cellval)
        book.save(outputFilePath)
    
    def updated_ms_excel_file(self,strFilePath,strsheetName,dctVarb):
        """ It retuen the list of registration codes"""
        try:
            exlObj = Dispatch("Excel.Application")
            exlObj.Application.Visible=False
            workbook = exlObj.Workbooks.Open(strFilePath)
            worksheet = workbook.Worksheets(strsheetName)
            colNames=[]
            used = worksheet.UsedRange
            
            intRowsCount =used.Row+used.Rows.Count-1
            #print "intRowsCount: "+str(intRowsCount)
            intColCount =used.Column + used.Columns.Count - 1
            #print "intColCount: "+str(intColCount)
            for iRowIndex in range(1,intRowsCount+1):
              for iColIndex in range(1,intColCount+1):
                cellValue = worksheet.Cells(iRowIndex,iColIndex).Value
                cellValue = str(cellValue)
                if iRowIndex==1:
                  colNames.append(cellValue)
                  continue
                if cellValue!=dctVarb['RecordNumber']:
                  continue
                worksheet.Cells(iRowIndex,int(colNames.index("Status"))+1).Value = dctVarb['Status']
                worksheet.Cells(iRowIndex,int(colNames.index("Message"))+1).Value = dctVarb['Message']
                worksheet.Cells(iRowIndex,int(colNames.index("ScreenShot"))+1).Value = dctVarb['ScreenShot']
            exlObj.ActiveSheet.Columns.AutoFit()
            workbook.Save()
            workbook.close
            exlObj.Application.Quit()
        except Exception as exp:
          print exp
          try:
            workbook.Save()
            workbook.close
            exlObj.Application.Quit()
          except:
            print "exp"
    
    def read_multiple_testdata(self,filepath,sheetname,testcasename):
        """read multiple rows of testdata based on testcase name"""
        try:
            workbook = xlrd.open_workbook(filepath)
            worksheet = workbook.sheet_by_name(sheetname)
            noofrows = worksheet.nrows
            print "noofrows: "+ str(noofrows)
            dictvar={}
            index=1
            for rowno in range(0,noofrows):
                cellvalue = worksheet.cell_value(rowno,0)
                rowValues = worksheet.row_values(rowno)
                if cellvalue == testcasename:
                    tempdict = {}
                    for colno in range(0,len(rowValues)):
                        keydata = worksheet.cell_value(0,colno)
                        celdata = worksheet.cell_value(rowno,colno)
                        if len(str(keydata))==0:
                            continue
                        if len(str(celdata))==0:
                            celdata = ""
                        tempdict[keydata] = celdata
                    dictvar[str(index)] = tempdict
                    index+=1
            return dictvar
        except Exception as exp:
            print "Got exception in read_multiple_testdata keyword.Error: "+str(exp)
            return {}


    def read_all_testdata(self,filepath,sheetname):
        """read multiple rows of testdata based on testcase name"""
        try:
            workbook = xlrd.open_workbook(filepath)
            worksheet = workbook.sheet_by_name(sheetname)
            noofrows = worksheet.nrows
            print "noofrows: "+ str(noofrows)
            dictvar={}
            for rowno in range(1,noofrows):
                rowno = int(rowno)
                cellvalue = worksheet.cell_value(rowno,0)
                rowValues = worksheet.row_values(rowno)
                tempdict = {}
                for colno in range(0,len(rowValues)):
                    keydata = worksheet.cell_value(0,colno)
                    celdata = worksheet.cell_value(rowno,colno)
                    if len(str(keydata))==0:
                        continue
                    if len(str(celdata))==0:
                        celdata = ""
                    tempdict[keydata] = celdata
                dictvar[str(index)] = tempdict 
            return dictvar
        except Exception as exp:
            print "Got exception in read_all_testdata keyword.Error: "+str(exp)
            return {}
        
    def run_excel_macro(self, sExcelFilePath, sMacroName):
        try:
             xlApp = win32com.client.DispatchEx('Excel.Application')
             xlsPath = os.path.expanduser(sExcelFilePath)
             wb = xlApp.Workbooks.Open(Filename=xlsPath)
             xlApp.Run(sMacroName)
             wb.Save()
             xlApp.Quit()
             print("Macro ran successfully!")
             root = Tk()
             prompt = 'Microsoft Excel'
             label1 = Label(root, text=prompt, width=len(prompt))
             label1.pack()
             root.destroy()
             root.mainloop()
             return True
        except Exception as exp:
            print("Error found while running the excel macro!"+str(exp))
            xlApp.Quit()
            return False
    def dialog():
        win = Toplevel()             
        Label(win,  text='Upload Files are created in the parent directory of excel file.').pack() 
        Button(win, text='OK', command=win.quit).pack()    
        win.protocol('WM_DELETE_WINDOW', win.quit)         
        win.focus_set()          
        win.grab_set()           
        win.mainloop()           
        win.destroy()
        print 'dialog exit' 
        root = Tk()
        Button(root, text='popup', command=dialog).pack()
        root.mainloop()

    def close_after_2s():
             root.destroy()
             root.after(2000, close_after_2s)
             root.mainloop()

    

#a=CustomLibrary()
#outputfileName='C:\\FG\\ErrorLogs\\SampleErrorLog.xlsx'
#outputfileName='C:\\FG\\SampleErrorLog.xlsx'


#ResultsDict={'1':{'BatchName':'sample', 'BatchStatus':'True', 'BatchDetails':'11367'}}
#a.create_result_log_file(outputfileName,ResultsDict)
